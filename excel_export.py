"""
Export Bookeo bookings to an Excel spreadsheet.
Flattens nested booking/customer/price data into rows.
"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Column order and header names for the Excel sheet
EXCEL_COLUMNS = [
    ("booking_number", "Booking #"),
    ("start_time", "Start Time"),
    ("end_time", "End Time"),
    ("title", "Title"),
    ("product_name", "Product"),
    ("product_id", "Product ID"),
    ("customer_id", "Customer ID"),
    ("customer_name", "Customer Name"),
    ("customer_email", "Customer Email"),
    ("customer_phone", "Customer Phone"),
    ("canceled", "Canceled"),
    ("cancelation_time", "Cancelation Time"),
    ("creation_time", "Created"),
    ("creation_agent", "Created By"),
    ("last_change_time", "Last Updated"),
    ("last_change_agent", "Last Updated By"),
    ("total_gross", "Total (Gross)"),
    ("total_net", "Total (Net)"),
    ("total_paid", "Total Paid"),
    ("currency", "Currency"),
    ("external_ref", "External Ref"),
    ("source", "Source"),
    ("no_show", "No Show"),
    ("resources", "Resources"),
    ("options", "Options"),
]


def _safe(val: Any, default: str = "") -> str:
    """Convert value to string for Excel; empty or None -> default."""
    if val is None:
        return default
    if isinstance(val, bool):
        return "Yes" if val else "No"
    if isinstance(val, (list, dict)):
        return str(val) if val else default
    return str(val).strip() or default


def _customer_name(customer: dict | None) -> str:
    if not customer:
        return ""
    parts = [
        customer.get("firstName"),
        customer.get("middleName"),
        customer.get("lastName"),
    ]
    return " ".join(p for p in parts if p).strip()


def _customer_phone(customer: dict | None) -> str:
    if not customer:
        return ""
    phones = customer.get("phoneNumbers") or []
    if not phones:
        return ""
    return (phones[0].get("number") or "").strip()


def _resources_list(booking: dict) -> str:
    resources = booking.get("resources") or []
    if not resources:
        return ""
    names = [r.get("name") or r.get("id") or "" for r in resources]
    return ", ".join(n for n in names if n)


def _options_list(booking: dict) -> str:
    options = booking.get("options") or []
    if not options:
        return ""
    parts = []
    for o in options:
        name = o.get("name") or o.get("id") or ""
        value = o.get("value") or ""
        if name or value:
            parts.append(f"{name}: {value}" if name else value)
    return "; ".join(parts)


def _booking_to_row(booking: dict) -> dict[str, str]:
    """Convert one Bookeo booking object to a flat dict of column key -> value."""
    customer = booking.get("customer") or {}
    price = booking.get("price") or {}
    total_gross = price.get("totalGross") or {}
    total_net = price.get("totalNet") or {}
    total_paid = price.get("totalPaid") or {}
    currency = (total_gross or total_net or total_paid).get("currency", "")

    return {
        "booking_number": _safe(booking.get("bookingNumber")),
        "start_time": _safe(booking.get("startTime")),
        "end_time": _safe(booking.get("endTime")),
        "title": _safe(booking.get("title")),
        "product_name": _safe(booking.get("productName")),
        "product_id": _safe(booking.get("productId")),
        "customer_id": _safe(booking.get("customerId")),
        "customer_name": _customer_name(customer),
        "customer_email": _safe(customer.get("emailAddress")),
        "customer_phone": _customer_phone(customer),
        "canceled": "Yes" if booking.get("canceled") else "No",
        "cancelation_time": _safe(booking.get("cancelationTime")),
        "creation_time": _safe(booking.get("creationTime")),
        "creation_agent": _safe(booking.get("creationAgent")),
        "last_change_time": _safe(booking.get("lastChangeTime")),
        "last_change_agent": _safe(booking.get("lastChangeAgent")),
        "total_gross": _safe(total_gross.get("amount")),
        "total_net": _safe(total_net.get("amount")),
        "total_paid": _safe(total_paid.get("amount")),
        "currency": _safe(currency),
        "external_ref": _safe(booking.get("externalRef")),
        "source": _safe(booking.get("source")),
        "no_show": "Yes" if booking.get("noShow") else "No",
        "resources": _resources_list(booking),
        "options": _options_list(booking),
    }


def write_bookings_to_excel(
    bookings: list[dict],
    output_path: str | Path,
    sheet_name: str = "Bookings",
) -> Path:
    """
    Write a list of Bookeo booking dicts to an Excel file.
    Uses the first row as headers (bold). Returns the output path.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel sheet name limit

    # Headers
    headers = [label for _key, label in EXCEL_COLUMNS]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    # Rows (dedupe by booking_number so we keep latest)
    key_to_row: dict[str, dict[str, str]] = {}
    for b in bookings:
        row = _booking_to_row(b)
        bn = row["booking_number"]
        key_to_row[bn] = row  # last occurrence wins if duplicates

    sorted_rows = list(key_to_row.values())
    for row_idx, row in enumerate(sorted_rows, 2):
        for col_idx, (key, _) in enumerate(EXCEL_COLUMNS, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))

    # Auto-fit would require openpyxl column width iteration
    for col in range(1, len(EXCEL_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    wb.save(output_path)
    logger.info("Wrote %d bookings to %s", len(sorted_rows), output_path)
    return output_path
